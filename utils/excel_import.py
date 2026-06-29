import io
import json
from datetime import date, datetime

from openpyxl import load_workbook

# ── 공통 ──────────────────────────────────────────────────

def _cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_date(value):
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()

    s = _cell_str(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _parse_int(value, default=None, min_val=None, max_val=None):
    if value is None or value == "":
        return default
    try:
        n = int(float(value))
    except (TypeError, ValueError):
        return None
    if min_val is not None:
        n = max(min_val, n)
    if max_val is not None:
        n = min(max_val, n)
    return n


def _load_sheet_rows(file_bytes):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return None, [], [{"row": 0, "message": f"Excel 파일을 읽을 수 없습니다: {e}"}]

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return None, [], [{"row": 0, "message": "빈 파일입니다."}]

    return wb, [header_row, rows_iter], []


def _build_col_map(header_row, header_map):
    col_map = {}
    for i, header in enumerate(header_row):
        if header is None:
            continue
        key = _cell_str(header)
        if key in header_map:
            col_map[header_map[key]] = i
    return col_map


def _row_getter(row, col_map):
    def get(field):
        if field not in col_map:
            return None
        idx = col_map[field]
        return row[idx] if idx < len(row) else None

    return get


def _is_empty_row(row):
    return not row or all(c is None or _cell_str(c) == "" for c in row)


# ── History ───────────────────────────────────────────────

VALID_HISTORY_TYPES = {"메신저", "이메일", "구두/회의", "기타"}

HISTORY_HEADER_MAP = {
    "유형": "type",
    "type": "type",
    "날짜": "date",
    "date": "date",
    "제목": "title",
    "title": "title",
    "내용": "content",
    "content": "content",
    "키워드": "keywords",
    "keywords": "keywords",
}


def _parse_keywords(value):
    if value is None or value == "":
        return "[]"

    if isinstance(value, str):
        s = value.strip()
        if s.startswith("["):
            try:
                parsed = json.loads(s)
                if isinstance(parsed, list):
                    return json.dumps(parsed, ensure_ascii=False)
            except json.JSONDecodeError:
                pass
        kws = [k.strip() for k in s.replace("，", ",").split(",") if k.strip()]
        return json.dumps(kws, ensure_ascii=False)

    return "[]"


def _parse_history_type(value):
    type_ = _cell_str(value) or "기타"
    if type_ not in VALID_HISTORY_TYPES:
        return None, f"유형 '{type_}'은(는) 지원하지 않습니다."
    return type_, None


def parse_history_excel(file_bytes):
    wb, parts, errors = _load_sheet_rows(file_bytes)
    if wb is None:
        return [], errors

    header_row, rows_iter = parts
    col_map = _build_col_map(header_row, HISTORY_HEADER_MAP)

    if "content" not in col_map:
        wb.close()
        return [], [{
            "row": 1,
            "message": "필수 열 '내용'을 찾을 수 없습니다. (유형, 날짜, 제목, 내용, 키워드)",
        }]

    records = []
    for row_num, row in enumerate(rows_iter, start=2):
        if _is_empty_row(row):
            continue

        get = _row_getter(row, col_map)

        content = _cell_str(get("content"))
        if not content:
            errors.append({"row": row_num, "message": "내용이 비어 있습니다."})
            continue

        type_, type_err = _parse_history_type(get("type"))
        if type_err:
            errors.append({"row": row_num, "message": type_err})
            continue

        parsed_date = _parse_date(get("date"))
        if not parsed_date:
            errors.append({"row": row_num, "message": "날짜 형식이 올바르지 않습니다."})
            continue

        records.append({
            "type": type_,
            "date": parsed_date,
            "title": _cell_str(get("title")),
            "content": content,
            "keywords": _parse_keywords(get("keywords")),
        })

    wb.close()
    return records, errors


# ── Work ──────────────────────────────────────────────────

VALID_WORK_STATUSES = {"done", "ing", "issue", "plan"}

WORK_STATUS_MAP = {
    "done": "done",
    "ing": "ing",
    "issue": "issue",
    "plan": "plan",
    "완료": "done",
    "진행중": "ing",
    "이슈/리스크": "issue",
    "이슈": "issue",
    "리스크": "issue",
    "다음주 계획": "plan",
    "계획": "plan",
}

WORK_HEADER_MAP = {
    "업무명": "title",
    "title": "title",
    "상태": "status",
    "status": "status",
    "시작일": "start_date",
    "start_date": "start_date",
    "종료일": "end_date",
    "종료예정일": "end_date",
    "end_date": "end_date",
    "진행률": "progress",
    "progress": "progress",
    "메모": "description",
    "description": "description",
}


def _parse_work_status(value):
    raw = _cell_str(value).lower() if isinstance(value, str) else _cell_str(value)
    if not raw:
        return "ing", None
    status = WORK_STATUS_MAP.get(raw) or WORK_STATUS_MAP.get(_cell_str(value))
    if not status:
        return None, f"상태 '{_cell_str(value)}'은(는) 지원하지 않습니다."
    return status, None


def parse_work_excel(file_bytes):
    wb, parts, errors = _load_sheet_rows(file_bytes)
    if wb is None:
        return [], errors

    header_row, rows_iter = parts
    col_map = _build_col_map(header_row, WORK_HEADER_MAP)

    if "title" not in col_map:
        wb.close()
        return [], [{
            "row": 1,
            "message": "필수 열 '업무명'을 찾을 수 없습니다. (업무명, 상태, 시작일, 종료일, 진행률, 메모)",
        }]

    records = []
    for row_num, row in enumerate(rows_iter, start=2):
        if _is_empty_row(row):
            continue

        get = _row_getter(row, col_map)

        title = _cell_str(get("title"))
        if not title:
            errors.append({"row": row_num, "message": "업무명이 비어 있습니다."})
            continue

        status, status_err = _parse_work_status(get("status"))
        if status_err:
            errors.append({"row": row_num, "message": status_err})
            continue

        start_date = _parse_date(get("start_date"))
        if not start_date:
            errors.append({"row": row_num, "message": "시작일 형식이 올바르지 않습니다."})
            continue

        end_date = _parse_date(get("end_date"))
        if not end_date:
            errors.append({"row": row_num, "message": "종료일 형식이 올바르지 않습니다."})
            continue

        progress = _parse_int(get("progress"), default=0, min_val=0, max_val=100)
        if progress is None:
            errors.append({"row": row_num, "message": "진행률은 0~100 사이 숫자여야 합니다."})
            continue

        records.append({
            "title": title,
            "status": status,
            "start_date": start_date,
            "end_date": end_date,
            "progress": progress,
            "description": _cell_str(get("description")),
        })

    wb.close()
    return records, errors
